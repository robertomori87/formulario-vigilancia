# src/aprovacao_lta/routes.py

from flask import Blueprint, render_template, redirect, request, url_for, session, jsonify
from src.common.database import SessionLocal, Checklist
from sqlalchemy import select
from sqlalchemy.exc import SQLAlchemyError
import subprocess
import openpyxl
import os

checklist_bp = Blueprint(
    "checklist_bp",
    __name__,
    template_folder="templates"
)

@checklist_bp.route("/importar_excel")
def importar_excel_checklist():
    if not session.get("logado"):
        return redirect(url_for("receituario_bp.login"))

    try:
        result = subprocess.run(
            ["python", "importar_perguntas_excel.py"],
            capture_output=True,
            text=True,
            check=True
        )
        return jsonify({"status": "sucesso", "output": result.stdout}), 200
    except subprocess.CalledProcessError as e:
        return jsonify({"status": "erro", "mensagem": e.stderr}), 500


@checklist_bp.route("/checklist")
def exibir_checklist_projeto():
    db = SessionLocal()
    try:
        perguntas = db.execute(select(Checklist).order_by(Checklist.numero_item)).scalars().all()
    except Exception as e:
        return f"Erro ao buscar checklist: {e}", 500
    finally:
        db.close()
    return render_template("checklist_form.html", perguntas=perguntas)

@checklist_bp.route("/salvar_checklist", methods=["POST"])
def salvar_checklist_projeto():
    if not session.get("logado"):
        return redirect(url_for("receituario_bp.login"))

    db = SessionLocal()
    try:
        perguntas_ids_db = db.execute(select(Checklist.id)).scalars().all()

        for pergunta_id_db in perguntas_ids_db:
            resposta = request.form.get(f"resposta_{pergunta_id_db}")
            atende = request.form.get(f"atende_{pergunta_id_db}") == "true"
            nao_atende = request.form.get(f"nao_atende_{pergunta_id_db}") == "true"
            medida_mitigatoria = request.form.get(f"medida_mitigatoria_{pergunta_id_db}")
            # REMOVIDO: justificativa_nao_se_aplica = request.form.get(f"justificativa_nao_se_aplica_{pergunta_id_db}")
            
            comentario_requerente = request.form.get(f"comentario_requerente_{pergunta_id_db}")
            comentario_avaliador = request.form.get(f"comentario_avaliador_{pergunta_id_db}")

            pergunta_existente = db.execute(
                select(Checklist).filter_by(id=pergunta_id_db)
            ).scalar_one_or_none()

            if pergunta_existente:
                pergunta_existente.resposta = resposta
                pergunta_existente.atende = atende
                pergunta_existente.nao_atende = nao_atende
                pergunta_existente.medida_mitigatoria = medida_mitigatoria if medida_mitigatoria else None
                # REMOVIDO: pergunta_existente.justificativa_nao_se_aplica = justificativa_nao_se_aplica if justificativa_nao_se_aplica else None
                
                pergunta_existente.comentario_requerente = comentario_requerente if comentario_requerente else None
                pergunta_existente.comentario_avaliador = comentario_avaliador if comentario_avaliador else None
            else:
                print(f"Aviso: Pergunta com ID {pergunta_id_db} não encontrada para atualização.")

        db.commit()
        print("Checklist de projeto salvo com sucesso!")
        return redirect(url_for('checklist_bp.exibir_checklist_projeto'))

    except SQLAlchemyError as e:
        db.rollback()
        print(f"Erro SQLAlchemy ao salvar checklist de projeto: {e}")
        return f"Erro ao salvar o checklist de projeto no banco de dados. Detalhes: {str(e)}", 500
    except Exception as ex:
        db.rollback()
        print(f"Erro inesperado ao salvar checklist de projeto: {ex}")
        return f"Erro inesperado ao salvar o checklist de projeto: {str(ex)}", 500
    finally:
        db.close()


def atualizar_excel_com_respostas():
    caminho = os.path.join("src", "aprovacao_lta", "dados", "checklist_perguntas.xlsx")

    db = SessionLocal()
    try:
        checklist = db.query(Checklist).order_by(Checklist.numero_item).all()
        wb = openpyxl.load_workbook(caminho)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        col_indices = {
            "resposta": headers.index("Resposta") + 1,
            "atende": headers.index("Atende") + 1,
            "nao_atende": headers.index("Não Atende") + 1,
            "medida_mitigatoria": headers.index("Mitigação") + 1,
            "justificativa_nao_se_aplica": headers.index("Justificativa") + 1,
            "comentario_requerente": headers.index("Comentário Requerente") + 1,
            "comentario_avaliador": headers.index("Comentário Avaliador") + 1
        }

        for row in ws.iter_rows(min_row=2):
            numero_item = row[0].value
            if numero_item is None:
                continue

            pergunta = next((p for p in checklist if p.numero_item == int(numero_item)), None)
            if not pergunta:
                continue

            row[col_indices["resposta"] - 1].value = pergunta.resposta
            row[col_indices["atende"] - 1].value = "SIM" if pergunta.atende else ""
            row[col_indices["nao_atende"] - 1].value = "SIM" if pergunta.nao_atende else ""
            row[col_indices["medida_mitigatoria"] - 1].value = pergunta.medida_mitigatoria
            row[col_indices["justificativa_nao_se_aplica"] - 1].value = pergunta.justificativa_nao_se_aplica
            row[col_indices["comentario_requerente"] - 1].value = pergunta.comentario_requerente
            row[col_indices["comentario_avaliador"] - 1].value = pergunta.comentario_avaliador

        wb.save(caminho)
        print(f"[✔️] Checklist atualizado no Excel: {caminho}")

    except Exception as e:
        print(f"[ERRO] ao atualizar Excel: {e}")
    finally:
        db.close()
